from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb=Workbook()
ws=wb.active
ws.title='MySheet'

ws2=wb.create_sheet('Another',0)
print(wb.sheetnames)

wb.save('hello-openpyxl.xlsx')

wb2=load_workbook('regions.xlsx')
active_sheet=wb2.active
active_sheet['A1']=0

wb2.save('regions-modified.xlsx')