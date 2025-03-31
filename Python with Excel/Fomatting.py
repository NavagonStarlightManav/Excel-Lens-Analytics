from openpyxl.styles import Font, Alignment, numbers
from openpyxl import Workbook

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Assigning values to specific cells
cell1 = ws['A1']
cell2 = ws['B2']
cell3 = ws.cell(row=3, column=3)

cell1.value = 'Hello'
cell2.value = 0.5
cell3.value = 'World!'

# Formatting cells
cell1.font = Font(name='Georgia', size=16, color='CF3338')
cell2.number_format = numbers.FORMAT_PERCENTAGE
cell3.alignment = Alignment(vertical='bottom')

# Adjust column width and row height
ws.column_dimensions['A'].width = 20
ws.row_dimensions[3].height = 70

# Save the workbook
wb.save("styled_excel.xlsx")
