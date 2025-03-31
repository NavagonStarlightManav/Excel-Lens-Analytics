from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

my_cell = ws['A1']
print(my_cell)

print(my_cell.value)

cell_range=ws[1]
print(cell_range)

cell_range=ws[1:3]
print(cell_range)

col_c=ws['C']
print(col_c)

col_A_C=ws['A:C']
print(col_A_C)

from openpyxl.utils import get_column_letter
print(ws[get_column_letter(3)])

for rows in ws.iter_rows(min_row=1,min_col=1,max_row=2,max_col=3):
    for cell in rows:
        print(cell.value)