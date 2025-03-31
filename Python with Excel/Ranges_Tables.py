import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
wb = Workbook()

import seaborn as sns
penguins = sns.load_dataset('penguins')
penguins = penguins.sample(frac=.05, random_state=1234)
penguins.head()

species=penguins['species'].unique().tolist()
print(species)

ws1=wb.create_sheet('species')
ws1.append(species)
new_range=openpyxl.workbook.defined_name.DefinedName('species',attr_text='species!$A$1:$C$1')

wb.defined_names.add(new_range)

# DefinedName('species', attr_text='species!$A$1:$C$1') creates a named range 'species' that refers to the range A1:C1 in the "species" worksheet
#
# wb.defined_names.append(new_range) adds this named range to the workbook

ws2 = wb.create_sheet('penguins')

for r in dataframe_to_rows(penguins, index=False, header=True):
    ws2.append(r)

style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)

table = Table(displayName="penguins", ref="A1:" + get_column_letter(ws2.max_column) + str(ws2.max_row))

table.tableStyleInfo = style

ws2.add_table(table)
wb.save('ranges-tables.xlsx')