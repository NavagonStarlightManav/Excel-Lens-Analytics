import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font


# Read the CSV file into a DataFrame, specifying data types for some columns
df = pd.read_csv('crime.csv', encoding='utf-8', dtype={
    'INCIDENT_NUMBER': str,
    'OFFENSE_CODE': str,
    'OFFENSE_CODE_GROUP': str,
    'OFFENSE_DESCRIPTION': str,
    'DISTRICT': str,
    'REPORTING_AREA': str,
    'YEAR': str,
    'MONTH': str,
    'DAY_OF_WEEK': str,
    'HOUR': str
})

# Filter the DataFrame to include only rows where 'OFFENSE_CODE_GROUP' is 'Counterfeiting'
df1 = df.query('OFFENSE_CODE_GROUP == "Counterfeiting"')

# Calculate the total number of crimes and the number of counterfeiting crimes
total_crimes = len(df.index)
counterfeit = len(df1.index)

# Calculate the percentage of counterfeiting crimes and round it to 2 decimal places
perc_crimes = (counterfeit / total_crimes) * 100
perc_crimes = round(perc_crimes, 2)


print(perc_crimes)

# Assuming 'ws' is your active worksheet object in an openpyxl workbook
# If you don't have it, create a new workbook and get the active sheet:
wb = Workbook()
ws = wb.active

# Write the calculated values to cells in the worksheet
ws['O8'].value = total_crimes
ws['P8'].value = counterfeit
ws['Q8'].value = perc_crimes

df1 = df1.copy()
df1['count']=1

# initially at end of each crime count of crimes is equal to 1

df2 = df1.groupby(['DISTRICT', 'YEAR']).size().to_frame(name='count').unstack(level=0)

# group(['DISTRICT', 'YEAR']) → Groups the data based on DISTRICT and YEAR

# size() → Counts the number of crimes in each group

# to_frame(name='count') → Converts the result into a DataFrame with the column name count

# This reshapes the table so that DISTRICT names become column headers, and YEAR remains as the row index

# This makes it easier to export the data into Excel

# Transformed Table Example

# YEAR	A1	B2
# 2018	5	3
# 2019	8	6
# Now each DISTRICT is a separate column

ws.cell(row=7, column=1, value="YEAR")  # Column A

# YEAR	count (A1)	count (B2)	count (C3)
# 2018	   5	           8	   3
# 2019	   2	           7	   1

# for col_idx, district in enumerate(df2.columns, 2)

# but we want only first index values like A1,B2.....

for col_idx, district in enumerate(df2.columns.get_level_values(1), 2):
    ws.cell(row=7, column=col_idx, value=district)

rows = df2.itertuples(index=True, name=None)

# itertuples() converts each row of df2 into a tuple

# index=True keeps the YEAR in the tuple

# name=None removes the default tuple name


# Write the rows to the worksheet, starting from row 8 and column 1
for i,j in enumerate(rows, 8):
    for c_idx, value in enumerate(j, 1):
        ws.cell(row=i, column=c_idx, value=value)

# it is just a normal for loop like for j in rows , i is an index


# data went from crime.xlsx to df1 to df2 to rows to ws

wb.save('crime_report.xlsx')

# If df2 looks like this:

# YEAR	A1	B2
# 2018	5	3
# 2019	8	6

# Row 8, Column 1 → 2018
#
# Row 8, Column 2 → 5
#
# Row 8, Column 3 → 3
#
# Row 9, Column 1 → 2019
#
# Row 9, Column 2 → 8
#
# Row 9, Column 3 → 6
#
# So, Excel will have:

#      A      B       C
# 8  2018     5       3
# 9  2019     8       6

# Then, r_idx will take these values:

# r_idx = 8 → for (2018, 5, 3)
# r_idx = 9 → for (2019, 8, 6)
# So the first row of data goes into Excel row 8, the next into row 9, and so on.

# For the first row (2018, 5, 3), c_idx takes

# c_idx = 1 → 2018 goes into column A
# c_idx = 2 → 5 goes into column B
# c_idx = 3 → 3 goes into column C
