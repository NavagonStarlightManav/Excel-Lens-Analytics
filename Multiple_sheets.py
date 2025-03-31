import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')

# Concatenate the DataFrames vertically (along rows)
df_all = pd.concat([df_1, df_2, df_3], sort=False)

# Display the resulting DataFrame
print(df_all)

print(df_all.shape[0]==df_1.shape[0]+df_2.shape[0]+df_3.shape[0])

df_all.to_excel('all_shifts.xlsx',index=False)

wb=load_workbook('all_shifts.xlsx')
ws=wb.active

total_col = 'G1'  # Column G, row 1
ws[total_col].font = Font(bold=True)
ws[total_col].value = 'Total'


# Define the columns to multiply (E and F)
e_col, f_col = 'E', 'F'

# Loop through rows 2 to 299 (inclusive)
for row in range(2, 300):
    # Construct the cell reference for the 'Total' column in the current row
    result_cell = 'G{}'.format(row)

    # Get the values from columns E and F in the current row
    e_value = ws[e_col + str(row)].value
    f_value = ws[f_col + str(row)].value

    # Multiply the values and write the result to the 'Total' column
    ws[result_cell].value = e_value * f_value

# Save the workbook to a new Excel file
wb.save('totaled.xlsx')